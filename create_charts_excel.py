from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

#loading workbook
wb = load_workbook('C:\Coding\scripts\Generated\\beverage_sales_pivot_table.xlsx')
#selecting the sheet
sheet = wb['Report']

#gives the max and min values of the col and rows on the sheet
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

barchart = BarChart()

# data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row)
# categories = data = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row)

#generated chart only using 2 rows of data so modified the rows and columns accordingly..
#else use the commented code above one to get the whole data.

data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=3)
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=3)

#add data and categories to the barchart
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

#adds chart to the sheet, on specific location.. e.g here added on cell B25
sheet.add_chart(barchart, "B25")

#edit title
barchart.title = "Sales by Company"
#set the color code of the barchart
barchart.style = 2

wb.save("C:\Coding\scripts\Generated\\bar_chart.xlsx")