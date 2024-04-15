from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#loading workbook
wb = load_workbook('C:\Coding\scripts\Generated\\beverage_sales_pivot_table.xlsx')
#selecting the sheet
sheet = wb['Report']

#gives the max and min values of the col and rows on the sheet
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#we will generate formulas to calculate total sales of company in all countries
#so essentially the sum of the whole row

for i in range(min_row+1, max_row+1):
    # letter = get_column_letter(i)
    sheet[f'O{i}'] = f'=SUM(B{i}:N{i})'
    # print(f'=SUM({letter}{min_column+1}:{letter}{max_column})')
    sheet[f'O{i}'].style = 'Currency'

wb.save("C:\Coding\scripts\Generated\Report.xlsx")